import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    """Базовый класс для экспорта в Excel"""

    @staticmethod
    def _style_headers(ws, headers):
        """Стилизация заголовков"""
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        font = Font(bold=True, color='FFFFFF', size=11)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = 20

    @classmethod
    def export_clients(cls, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Клиенты"

        headers = [
            'ID', 'Тип', 'Полное наименование', 'Краткое наименование',
            'ИНН', 'КПП', 'ОГРН', 'Фамилия', 'Имя', 'Отчество',
            'Дата рождения', 'Телефон', 'Email', 'Сайт', 'Страна', 'Город',
            'Адрес', 'Почтовый индекс', 'Менеджер', 'Статус', 'Источник',
            'Дата создания', 'Дата обновления'
        ]
        cls._style_headers(ws, headers)

        for row_idx, client in enumerate(queryset, 2):
            row_data = [
                client.id,
                client.get_type_display(),
                client.full_name,
                client.short_name or '',
                client.inn or '',
                client.kpp or '',
                client.ogrn or '',
                client.last_name or '',
                client.first_name or '',
                client.patronymic or '',
                client.birth_date.strftime('%Y-%m-%d') if client.birth_date else '',
                client.phone or '',
                client.email or '',
                client.website or '',
                client.country or '',
                client.city or '',
                client.address or '',
                client.postal_code or '',
                str(client.manager) if client.manager else '',
                client.get_status_display(),
                client.source or '',
                client.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                client.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def export_contracts(cls, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Контракты"

        headers = [
            'ID', 'Номер', 'Название', 'Клиент', 'Тип', 'Сумма', 'Оплачено',
            'Остаток', 'Процент оплаты', 'Статус оплаты', 'Статус',
            'Дата начала', 'Дата окончания', 'Дата подписания', 'Менеджер',
            'Дата создания', 'Дата обновления'
        ]
        cls._style_headers(ws, headers)

        for row_idx, contract in enumerate(queryset, 2):
            row_data = [
                contract.id,
                contract.number,
                contract.title,
                contract.client.full_name if contract.client else '',
                contract.get_type_display(),
                float(contract.amount),
                float(contract.paid_amount),
                float(contract.remaining_amount),
                round(contract.payment_percentage, 2),
                contract.get_payment_status_display(),
                contract.get_status_display(),
                contract.start_date.strftime('%Y-%m-%d') if contract.start_date else '',
                contract.end_date.strftime('%Y-%m-%d') if contract.end_date else '',
                contract.signed_date.strftime('%Y-%m-%d') if contract.signed_date else '',
                str(contract.manager) if contract.manager else '',
                contract.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                contract.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def export_tasks(cls, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Задачи"

        headers = [
            'ID', 'Название', 'Описание', 'Статус', 'Приоритет',
            'Срок выполнения', 'Дата завершения', 'Исполнитель', 'Клиент',
            'Контракт', 'Плановые часы', 'Фактические часы', 'Дата создания'
        ]
        cls._style_headers(ws, headers)

        for row_idx, task in enumerate(queryset, 2):
            row_data = [
                task.id,
                task.title,
                task.description,
                task.get_status_display(),
                task.get_priority_display(),
                task.due_date.strftime('%Y-%m-%d %H:%M:%S') if task.due_date else '',
                task.completed_at.strftime('%Y-%m-%d %H:%M:%S') if task.completed_at else '',
                str(task.assigned_to) if task.assigned_to else '',
                task.client.full_name if task.client else '',
                task.contract.number if task.contract else '',
                float(task.estimated_hours) if task.estimated_hours else '',
                float(task.actual_hours) if task.actual_hours else '',
                task.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    def export_payments(cls, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = "Платежи"

        headers = [
            'ID', 'Контракт', 'Сумма', 'Дата платежа', 'Способ оплаты',
            'Комментарий', 'Статус в ЮKassa', 'Дата фактической оплаты',
            'Кем добавлен', 'Дата создания'
        ]
        cls._style_headers(ws, headers)

        for row_idx, payment in enumerate(queryset, 2):
            row_data = [
                payment.id,
                payment.contract.number if payment.contract else '',
                float(payment.amount),
                payment.payment_date.strftime('%Y-%m-%d'),
                payment.get_payment_method_display(),
                payment.comment,
                payment.yookassa_status,
                payment.paid_at.strftime('%Y-%m-%d %H:%M:%S') if payment.paid_at else '',
                str(payment.created_by) if payment.created_by else '',
                payment.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            ]
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
